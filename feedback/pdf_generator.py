"""
PDF Generator Module
Generates PDF, Text and Excel reports for feedback export
"""

from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_JUSTIFY
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import os
from datetime import datetime


def create_pdf_report(data, output_path, teacher_name):
    """
    Create a PDF report from analysis data
    
    Args:
        data: Dictionary containing all analysis data
        output_path: Path to save PDF
        teacher_name: Name of the teacher
    
    Returns:
        Path to generated PDF
    """
    # Create document
    doc = SimpleDocTemplate(
        output_path,
        pagesize=A4,
        rightMargin=72,
        leftMargin=72,
        topMargin=72,
        bottomMargin=72
    )
    
    # Build content
    story = []
    styles = getSampleStyleSheet()
    
    # Custom styles
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#8b5cf6'),
        spaceAfter=30,
        alignment=TA_CENTER
    )
    
    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=14,
        textColor=colors.HexColor('#6366f1'),
        spaceAfter=12,
        spaceBefore=20
    )
    
    subheading_style = ParagraphStyle(
        'CustomSubHeading',
        parent=styles['Heading3'],
        fontSize=12,
        textColor=colors.HexColor('#06b6d4'),
        spaceAfter=10,
        spaceBefore=10
    )
    
    normal_style = ParagraphStyle(
        'CustomNormal',
        parent=styles['Normal'],
        fontSize=10,
        spaceAfter=8,
        alignment=TA_JUSTIFY
    )
    
    # Title
    story.append(Paragraph("Assignment Analytics Report", title_style))
    story.append(Spacer(1, 10))
    
    # Date and Teacher
    date_str = datetime.now().strftime("%B %d, %Y at %H:%M")
    story.append(Paragraph(f"<b>Generated:</b> {date_str}", normal_style))
    story.append(Paragraph(f"<b>Teacher:</b> {teacher_name}", normal_style))
    story.append(Spacer(1, 20))
    
    # Overall Summary
    if 'overall_summary' in data:
        story.append(Paragraph("Overall Summary", heading_style))
        summary = data['overall_summary']
        
        # Summary table
        summary_data = [
            ['Total Students', str(summary.get('total_students', 0))],
            ['Total Questions', str(summary.get('total_questions', 0))],
            ['Overall Similarity', f"{summary.get('overall_similarity', 0)}%"],
            ['Average Insight Score', f"{summary.get('avg_insight_score', 0)}"],
        ]
        
        summary_table = Table(summary_data, colWidths=[2.5*inch, 2*inch])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#f3f4f6')),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.HexColor('#374151')),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#e5e7eb'))
        ]))
        story.append(summary_table)
        story.append(Spacer(1, 20))
    
    # Question-wise Analysis
    if 'questions' in data:
        for q_id, q_data in data['questions'].items():
            story.append(Paragraph(f"Question {q_id}: {q_data.get('question_text', '')}", heading_style))
            
            # Question stats
            q_stats = [
                ['Total Responses', str(q_data.get('total_responses', 0))],
                ['Insight Score', str(q_data.get('insight_score', 0))],
                ['Confidence Score', str(q_data.get('confidence_score', 0))],
                ['Understanding Level', q_data.get('understanding_level', 'N/A')],
                ['Risk Level', q_data.get('risk_level', 'N/A')],
            ]
            
            q_table = Table(q_stats, colWidths=[2.5*inch, 2*inch])
            q_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#e0e7ff')),
                ('TEXTCOLOR', (0, 0), (-1, -1), colors.HexColor('#3730a3')),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
                ('TOPPADDING', (0, 0), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#c7d2fe'))
            ]))
            story.append(q_table)
            story.append(Spacer(1, 15))
            
            # Teaching Recommendation
            if 'teaching_action' in q_data:
                story.append(Paragraph("Teaching Recommendation:", subheading_style))
                story.append(Paragraph(q_data['teaching_action'], normal_style))
                story.append(Spacer(1, 10))
            
            # Common Keywords
            if 'common_keywords' in q_data and q_data['common_keywords']:
                story.append(Paragraph("Common Keywords:", subheading_style))
                keywords = ", ".join(q_data['common_keywords'])
                story.append(Paragraph(keywords, normal_style))
                story.append(Spacer(1, 10))
            
            # Weak Concepts
            if 'weak_concepts' in q_data:
                story.append(Paragraph("Areas Needing Attention:", subheading_style))
                for concept, value in q_data['weak_concepts'].items():
                    if isinstance(value, bool) and value:
                        concept_name = concept.replace('_', ' ').title()
                        story.append(Paragraph(f"• {concept_name}", normal_style))
                    elif isinstance(value, int) and value > 0:
                        story.append(Paragraph(f"• {value} short answers detected", normal_style))
                story.append(Spacer(1, 10))
            
            story.append(Spacer(1, 10))
    
    # Footer
    story.append(Spacer(1, 30))
    story.append(Paragraph(
        "<i>This report was automatically generated by AI Assignment Analytics System.</i>",
        ParagraphStyle('Footer', parent=styles['Normal'], fontSize=8, textColor=colors.grey, alignment=TA_CENTER)
    ))
    
    # Build PDF
    doc.build(story)
    
    return output_path


def generate_text_report(data, output_path, teacher_name):
    """
    Generate a plain text report
    
    Args:
        data: Dictionary containing all analysis data
        output_path: Path to save text file
        teacher_name: Name of the teacher
    
    Returns:
        Path to generated text file
    """
    lines = []
    
    # Header
    lines.append("=" * 60)
    lines.append("ASSIGNMENT ANALYTICS REPORT")
    lines.append("=" * 60)
    lines.append("")
    lines.append(f"Generated: {datetime.now().strftime('%B %d, %Y at %H:%M')}")
    lines.append(f"Teacher: {teacher_name}")
    lines.append("")
    
    # Overall Summary
    if 'overall_summary' in data:
        lines.append("-" * 60)
        lines.append("OVERALL SUMMARY")
        lines.append("-" * 60)
        summary = data['overall_summary']
        lines.append(f"Total Students: {summary.get('total_students', 0)}")
        lines.append(f"Total Questions: {summary.get('total_questions', 0)}")
        lines.append(f"Overall Similarity: {summary.get('overall_similarity', 0)}%")
        lines.append(f"Average Insight Score: {summary.get('avg_insight_score', 0)}")
        lines.append("")
    
    # Question-wise Analysis
    if 'questions' in data:
        for q_id, q_data in data['questions'].items():
            lines.append("-" * 60)
            lines.append(f"QUESTION {q_id}: {q_data.get('question_text', '')}")
            lines.append("-" * 60)
            lines.append(f"Total Responses: {q_data.get('total_responses', 0)}")
            lines.append(f"Insight Score: {q_data.get('insight_score', 0)}")
            lines.append(f"Confidence Score: {q_data.get('confidence_score', 0)}")
            lines.append(f"Understanding Level: {q_data.get('understanding_level', 'N/A')}")
            lines.append(f"Risk Level: {q_data.get('risk_level', 'N/A')}")
            lines.append("")
            
            if 'teaching_action' in q_data:
                lines.append("TEACHING RECOMMENDATION:")
                lines.append(q_data['teaching_action'])
                lines.append("")
            
            if 'common_keywords' in q_data and q_data['common_keywords']:
                lines.append("COMMON KEYWORDS:")
                lines.append(", ".join(q_data['common_keywords']))
                lines.append("")
            
            if 'weak_concepts' in q_data:
                lines.append("AREAS NEEDING ATTENTION:")
                for concept, value in q_data['weak_concepts'].items():
                    if isinstance(value, bool) and value:
                        lines.append(f"  - {concept.replace('_', ' ').title()}")
                    elif isinstance(value, int) and value > 0:
                        lines.append(f"  - {value} short answers detected")
                lines.append("")
    
    lines.append("=" * 60)
    lines.append("Generated by AI Assignment Analytics System")
    lines.append("=" * 60)
    
    # Write to file
    with open(output_path, 'w', encoding='utf-8') as f:
        f.write('\n'.join(lines))
    
    return output_path


def create_excel_report(data, output_path, teacher_name):
    """
    Generate an Excel report with multiple sheets
    
    Args:
        data: Dictionary containing all analysis data
        output_path: Path to save Excel file
        teacher_name: Name of the teacher
    
    Returns:
        Path to generated Excel file
    """
    # Create workbook
    wb = openpyxl.Workbook()
    
    # Styles
    header_fill = PatternFill(start_color="6366f1", end_color="6366f1", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # ============ Sheet 1: Summary ============
    ws_summary = wb.active
    ws_summary.title = "Summary"
    
    # Title
    ws_summary.merge_cells('A1:D1')
    ws_summary['A1'] = "Assignment Analytics Report"
    ws_summary['A1'].font = Font(bold=True, size=16, color="8b5cf6")
    ws_summary['A1'].alignment = Alignment(horizontal='center')
    
    # Date and Teacher
    ws_summary['A3'] = "Generated:"
    ws_summary['B3'] = datetime.now().strftime("%B %d, %Y at %H:%M")
    ws_summary['A4'] = "Teacher:"
    ws_summary['B4'] = teacher_name
    
    # Overall Summary
    ws_summary['A6'] = "OVERALL SUMMARY"
    ws_summary['A6'].font = Font(bold=True, size=14)
    
    if 'overall_summary' in data:
        summary = data['overall_summary']
        summary_headers = ['Metric', 'Value']
        summary_data = [
            ['Total Students', summary.get('total_students', 0)],
            ['Total Questions', summary.get('total_questions', 0)],
            ['Overall Similarity', f"{summary.get('overall_similarity', 0)}%"],
            ['Average Insight Score', summary.get('avg_insight_score', 0)],
        ]
        
        # Write headers
        for col, header in enumerate(summary_headers, 1):
            cell = ws_summary.cell(row=7, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
        
        # Write data
        for row_idx, row_data in enumerate(summary_data, 8):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws_summary.cell(row=row_idx, column=col_idx)
                cell.value = value
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center' if col_idx > 1 else 'left')
    
    # Adjust column widths
    ws_summary.column_dimensions['A'].width = 25
    ws_summary.column_dimensions['B'].width = 20
    
    # ============ Sheet 2: Question Analysis ============
    ws_questions = wb.create_sheet("Question Analysis")
    
    # Headers
    q_headers = ['Question ID', 'Question Text', 'Total Responses', 'Insight Score', 
                'Confidence Score', 'Understanding Level', 'Risk Level', 'Teaching Recommendation']
    
    for col, header in enumerate(q_headers, 1):
        cell = ws_questions.cell(row=1, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = thin_border
    
    # Data
    if 'questions' in data:
        for row_idx, (q_id, q_data) in enumerate(data['questions'].items(), 2):
            ws_questions.cell(row=row_idx, column=1).value = q_id
            ws_questions.cell(row=row_idx, column=2).value = q_data.get('question_text', '')
            ws_questions.cell(row=row_idx, column=3).value = q_data.get('total_responses', 0)
            ws_questions.cell(row=row_idx, column=4).value = q_data.get('insight_score', 0)
            ws_questions.cell(row=row_idx, column=5).value = q_data.get('confidence_score', 0)
            ws_questions.cell(row=row_idx, column=6).value = q_data.get('understanding_level', 'N/A')
            ws_questions.cell(row=row_idx, column=7).value = q_data.get('risk_level', 'N/A')
            ws_questions.cell(row=row_idx, column=8).value = q_data.get('teaching_action', '')
            
            # Apply borders
            for col in range(1, 9):
                ws_questions.cell(row=row_idx, column=col).border = thin_border
    
    # Adjust column widths
    ws_questions.column_dimensions['A'].width = 12
    ws_questions.column_dimensions['B'].width = 40
    ws_questions.column_dimensions['C'].width = 15
    ws_questions.column_dimensions['D'].width = 15
    ws_questions.column_dimensions['E'].width = 18
    ws_questions.column_dimensions['F'].width = 20
    ws_questions.column_dimensions['G'].width = 12
    ws_questions.column_dimensions['H'].width = 50
    
    # ============ Sheet 3: Keywords & Weak Concepts ============
    ws_keywords = wb.create_sheet("Keywords & Concepts")
    
    # Headers
    kw_headers = ['Question ID', 'Common Keywords', 'Short Answers Count', 'Low Vocabulary Diversity']
    
    for col, header in enumerate(kw_headers, 1):
        cell = ws_keywords.cell(row=1, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    
    # Data
    if 'questions' in data:
        for row_idx, (q_id, q_data) in enumerate(data['questions'].items(), 2):
            ws_keywords.cell(row=row_idx, column=1).value = q_id
            
            # Keywords
            keywords = ", ".join(q_data.get('common_keywords', []))
            ws_keywords.cell(row=row_idx, column=2).value = keywords
            
            # Weak concepts
            weak = q_data.get('weak_concepts', {})
            ws_keywords.cell(row=row_idx, column=3).value = weak.get('short_answers', 0)
            ws_keywords.cell(row=row_idx, column=4).value = "Yes" if weak.get('low_vocab_diversity', False) else "No"
            
            # Apply borders
            for col in range(1, 5):
                ws_keywords.cell(row=row_idx, column=col).border = thin_border
    
    # Adjust column widths
    ws_keywords.column_dimensions['A'].width = 12
    ws_keywords.column_dimensions['B'].width = 50
    ws_keywords.column_dimensions['C'].width = 18
    ws_keywords.column_dimensions['D'].width = 22
    
    # Save workbook
    wb.save(output_path)
    
    return output_path
